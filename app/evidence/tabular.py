"""Onda Tabular — ingestão de CSV/XLSX para tabela DuckDB consultável.

Três operações:

1. `analyze_tabular(data, filename, mime_type)` — dry-run em memória:
   carrega via DuckDB :memory:, infere schema, calcula score de qualidade,
   detecta PK candidata. NÃO persiste nada. Usado pelo modal
   "Promover para tabela?".

2. `promote_to_table(ks_id, data, filename, name, ...)` — cria arquivo
   .duckdb persistente em `data/tabular/<ks_id>/<table_id>.duckdb`,
   ingere os dados, grava metadata em Postgres (data_tables).
   Idempotente via versionamento por slug (re-promove gera v2, v3...).

3. `execute_query(table_id, inputs, select, filters, order_by, limit, ...)`
   — abre o .duckdb em modo READ-ONLY, valida select/filters contra o
   schema, monta SELECT parametrizado com bind vars (?), executa, audita
   em data_table_query_logs. Limit enforcement.

Decisões:
- Engine: DuckDB embarcado. read_csv_auto/read_xlsx nativos. Read-only
  por execução = safety técnica (não só prompt).
- Storage: 1 arquivo .duckdb por tabela (isolamento total, drop = rm).
  Caminho: `data/tabular/<ks_id>/<table_id>.duckdb`.
- Nome da tabela interna: sempre "data" (1 tabela por arquivo no MVP).
- Tipos: inferidos automaticamente pelo DuckDB (smart inference).
- Limites: ver `app/data_tables/types.py`.
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any, Optional

from app.core.config import get_settings
from app.core.database import (
    _get_pool,
    data_table_query_logs_repo,
    data_tables_repo,
    knowledge_repo,
)
from app.data_tables.events import (
    EVT_ANALYZE_COMPLETED,
    EVT_ANALYZE_FAILED,
    EVT_ANALYZE_STARTED,
    EVT_APPEND_COMPLETED,
    EVT_APPEND_FAILED,
    EVT_APPEND_STARTED,
    EVT_DUCKDB_ERROR,
    EVT_PROMOTE_COMPLETED,
    EVT_PROMOTE_FAILED,
    EVT_PROMOTE_STARTED,
    EVT_QUERY_EXECUTED,
)
from app.data_tables.queries import (
    build_urn,
    find_by_id_with_ks,
    next_version_for_slug,
    slugify,
)
from app.data_tables.types import (
    MAX_COLUMNS,
    MAX_ROWS_RETURNED,
    MAX_TABLE_SIZE_MB,
    SqlOperator,
    render_where_clause,
)

logger = logging.getLogger(__name__)
# Logger dedicado da Onda Tabular — escreve em logs/tabular.log
_tabular_logger = logging.getLogger("tabular")


class TabularError(Exception):
    """Erro de ingestão/query tabular com HTTP status sugerido."""

    def __init__(self, message: str, status_code: int = 500):
        super().__init__(message)
        self.status_code = status_code


# ─── Internals: DuckDB e arquivo ─────────────────────────────────


_DUCKDB_TABLE = "data"  # Nome da tabela interna no .duckdb (1 por arquivo)
_TABULAR_ROOT = Path("data") / "tabular"


def _import_duckdb():
    """Import lazy para não pagar custo de import quando feature não é usada."""
    try:
        import duckdb  # type: ignore
        return duckdb
    except ImportError as e:
        raise TabularError(
            "DuckDB não instalado. Execute `pip install duckdb>=1.0.0` "
            "ou `pip install -r requirements.txt`.",
            status_code=503,
        ) from e


def _ext_from_filename(filename: str) -> str:
    """Retorna 'csv' | 'xlsx' | 'unsupported' a partir da extensão."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return "csv"
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return "xlsx"
    return "unsupported"


def _list_xlsx_sheets(file_path: str) -> list[str]:
    """Lista nomes das abas via openpyxl (já presente como dep de markitdown[all]).

    Retorna `[]` se openpyxl indisponível ou arquivo corrompido — caller
    cai no fluxo de "1 aba" como fallback.
    """
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception as e:
        logger.debug("openpyxl list sheets falhou: %s", e)
        return []


def _xlsx_sheet_dimensions(file_path: str, sheet_name: str) -> Optional[tuple[int, int]]:
    """Retorna (max_row, max_col) da aba via openpyxl. None se falhar.

    Necessário para construir `range=` válido para DuckDB — o read_xlsx exige
    range no formato `A{r1}:{col_letter}{r2}` (não aceita `A2`, nem `ZZZ`).
    """
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return None
            ws = wb[sheet_name]
            return (ws.max_row or 0, ws.max_column or 0)
        finally:
            wb.close()
    except Exception as e:
        logger.debug("openpyxl dimensions falhou: %s", e)
        return None


def _col_letter(idx: int) -> str:
    """Converte índice 1-based em letra de coluna Excel (1='A', 27='AA')."""
    s = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s or "A"


def _xlsx_range_for_header(file_path: str, sheet_name: str, header_row: int) -> Optional[str]:
    """LEGACY: range para DuckDB `read_xlsx` (não mais usado — XLSX vai via
    openpyxl→CSV em `_xlsx_sheet_to_csv`). Mantido por retro-compatibilidade
    com testes existentes.
    """
    if header_row <= 1:
        return None
    if not sheet_name:
        return None
    dims = _xlsx_sheet_dimensions(file_path, sheet_name)
    if not dims:
        return None
    max_row, max_col = dims
    if max_row < header_row or max_col < 1:
        return None
    return f"A{header_row}:{_col_letter(max_col)}{max_row}"


def _xlsx_sheet_to_csv(
    xlsx_path: str,
    sheet_name: Optional[str],
    header_row: int,
    out_csv_path: str,
) -> None:
    """Converte UMA aba do XLSX para CSV via openpyxl.

    Por que: a extension `excel` do DuckDB requer download em runtime
    (`INSTALL excel; LOAD excel`) — quebra em ambientes corporativos sem
    internet ou sem permissão de instalar extensions. openpyxl já está
    presente (dep transitiva de `markitdown[all]`) e gera CSV que DuckDB
    lê nativamente via `read_csv_auto` (built-in, zero deps).

    Pula `header_row - 1` linhas iniciais. Ex: header_row=2 → descarta
    a linha 1 (provavelmente título mergeado) e usa linha 2 como header
    do CSV.

    Args:
        xlsx_path: caminho do XLSX origem.
        sheet_name: nome da aba. None = primeira aba (wb.active).
        header_row: linha (1-based) que vira a primeira linha do CSV.
        out_csv_path: caminho do CSV destino (sobrescreve).

    Raises:
        TabularError: openpyxl indisponível, XLSX corrompido, ou aba ausente.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise TabularError(
            "openpyxl indisponível — necessário para ler XLSX. "
            "Execute `pip install openpyxl` (ou `pip install -r requirements.txt`).",
            status_code=503,
        ) from e

    import csv
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        raise TabularError(
            f"Falha ao abrir XLSX: {e}. Arquivo pode estar corrompido ou "
            f"protegido por senha.",
            status_code=400,
        ) from e
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise TabularError(
                    f"Aba '{sheet_name}' não encontrada. Disponíveis: {wb.sheetnames}",
                    status_code=400,
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Escreve CSV. newline='' segue convenção do módulo csv (evita
        # linhas duplicadas no Windows). encoding utf-8 para acentos.
        # None → '' (CSV não tem null; DuckDB infere NULL pra string vazia
        # em colunas numéricas).
        with open(out_csv_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            skip = max(0, header_row - 1)
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i <= skip:
                    continue
                writer.writerow(["" if v is None else v for v in row])
    finally:
        wb.close()


def _read_into_duckdb(
    con,
    file_path: str,
    ext: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
) -> None:
    """Cria `data` table a partir do arquivo.

    CSV: usa `read_csv_auto` (built-in DuckDB, infere tipos + delimiter).
    XLSX: converte aba via `_xlsx_sheet_to_csv` (openpyxl) → CSV temp →
    `read_csv_auto`. NÃO depende da extension `excel` do DuckDB (que
    exige download em runtime e quebra em ambientes offline/corporativos).

    Args:
        con: conexão DuckDB ativa.
        file_path: caminho local do arquivo.
        ext: 'csv' ou 'xlsx'.
        sheet_name: nome da aba do XLSX. None = primeira aba.
        header_row: linha (1-based) que contém os headers. Default 1.
            Útil quando XLSX tem título mergeado na linha 1 e headers na 2.
    """
    if ext == "csv":
        # read_csv_auto: HEADER detect, type inference, delimiter sniffer.
        # SKIP=N pula N linhas iniciais (header fica na linha N+1).
        safe = file_path.replace("'", "''")
        skip = max(0, header_row - 1)
        skip_clause = f", SKIP={skip}" if skip > 0 else ""
        con.execute(
            f"CREATE TABLE {_DUCKDB_TABLE} AS "
            f"SELECT * FROM read_csv_auto('{safe}', HEADER=TRUE{skip_clause})"
        )
        return
    if ext == "xlsx":
        # Pipeline: openpyxl lê XLSX → escreve CSV temp → DuckDB read_csv_auto.
        # header_row já é tratado no momento da conversão (skip de linhas).
        # Sem dependência da extension 'excel' do DuckDB.
        tmp_handle = tempfile.NamedTemporaryFile(
            suffix=".csv", delete=False, mode="w", encoding="utf-8", newline=""
        )
        tmp_csv_path = tmp_handle.name
        tmp_handle.close()
        try:
            _xlsx_sheet_to_csv(file_path, sheet_name, header_row, tmp_csv_path)
            safe = tmp_csv_path.replace("'", "''")
            # CSV já vem com header na primeira linha (skip de linhas foi feito
            # no _xlsx_sheet_to_csv), então não precisa SKIP aqui.
            con.execute(
                f"CREATE TABLE {_DUCKDB_TABLE} AS "
                f"SELECT * FROM read_csv_auto('{safe}', HEADER=TRUE)"
            )
        finally:
            try:
                os.unlink(tmp_csv_path)
            except OSError:
                pass
        return
    raise TabularError(
        f"Formato não suportado: {ext}. Aceitos: csv, xlsx.",
        status_code=400,
    )


def _describe_schema(con) -> list[dict]:
    """DESCRIBE data → [{name, type, nullable}]."""
    rows = con.execute(f"DESCRIBE {_DUCKDB_TABLE}").fetchall()
    # DESCRIBE retorna: (column_name, column_type, null, key, default, extra)
    schema = []
    for r in rows:
        schema.append({
            "name": r[0],
            "type": r[1],
            "nullable": str(r[2]).upper() == "YES",
        })
    return schema


def _looks_like_table_title(name: str) -> bool:
    """Heurística: nome de coluna que mais parece nome de tabela que header.

    Exemplos: TB_ANALISE_CREDITO, TAB_VENDAS, FATO_PEDIDOS, DIM_CLIENTE.
    Padrões UPPER_SNAKE_CASE com prefixo TB_ / TAB_ / FATO_ / DIM_ /
    STAGE_ ou cobrindo a coluna inteira em maiúsculo com underscores.
    """
    n = str(name or "").strip()
    if not n:
        return False
    if "_" not in n:
        return False
    upper_only = n.replace("_", "").isupper() and any(c.isalpha() for c in n)
    if not upper_only:
        return False
    prefixes = ("TB_", "TAB_", "FATO_", "FACT_", "DIM_", "STAGE_", "STG_")
    return any(n.startswith(p) for p in prefixes) or len(n) >= 10


def _compute_quality_score(con, schema: list[dict], row_count: int) -> tuple[float, list[str], Optional[str]]:
    """Score (0.0-1.0) heurístico + warnings + PK sugerida.

    Penalidades:
    - Schema vazio: 0
    - Caso patológico "1 col VARCHAR + rows>5 + 100% única": -0.7 (forte)
      Sinal de XLSX com título mergeado na linha 1 mal-interpretado como header.
    - Nome da coluna parece título de tabela (TB_X, UPPER_SNAKE): -0.5
    - row_count < 2: -0.5 (provavelmente só header)
    - Colunas 100% nulas: -0.3
    - Colunas >50% nulos: -0.2
    - Colunas sem nome (column0, column1...): -0.2

    PK candidata: primeira coluna com COUNT(DISTINCT col) == COUNT(*) AND
    COUNT(*) WHERE col IS NULL == 0.
    """
    score = 1.0
    warnings: list[str] = []

    if not schema:
        return 0.0, ["Schema vazio."], None

    # Detecção do caso patológico ANTES das checagens normais —
    # tipicamente XLSX com header em merged cell na linha 1.
    if len(schema) == 1 and row_count > 5:
        only = schema[0]
        only_type = str(only.get("type") or "").upper()
        if "VARCHAR" in only_type or "TEXT" in only_type or only_type == "":
            try:
                qcol = '"' + only["name"].replace('"', '""') + '"'
                r = con.execute(
                    f"SELECT COUNT(*), COUNT(DISTINCT {qcol}) FROM {_DUCKDB_TABLE}"
                ).fetchone()
                total, distinct_v = r[0], r[1]
                # 100% único = cada linha é um valor diferente → não parece dado tabular
                if total > 0 and distinct_v == total:
                    score = 0.0
                    warning = (
                        f"Detectada 1 coluna VARCHAR única ('{only['name']}') com {row_count} "
                        f"valores distintos. Provavelmente a linha 1 do XLSX é um título "
                        f"mergeado e não o cabeçalho real. Tente promover de novo selecionando "
                        f"a linha 2 como header."
                    )
                    if _looks_like_table_title(only["name"]):
                        warning += f" O nome '{only['name']}' parece nome de tabela, reforçando essa hipótese."
                    warnings.append(warning)
                    # Retorna cedo: outras penalidades não ajudam, já está zero.
                    return round(score, 3), warnings, None
            except Exception:
                pass

    # Nome da coluna parece nome de tabela (TB_*, etc) — mesmo com várias cols
    title_like_cols = [c["name"] for c in schema if _looks_like_table_title(c["name"])]
    if title_like_cols:
        score -= 0.5
        warnings.append(
            f"Coluna(s) com nome de tabela (ex: {title_like_cols[0]}). "
            f"Provavelmente a linha 1 do XLSX é título — tente promover usando linha 2 como header."
        )

    if row_count < 2:
        score -= 0.5
        warnings.append(f"Apenas {row_count} linha(s) — planilha quase vazia.")

    # Colunas sem nome real (DuckDB usa "column0" etc quando não há header)
    generic_cols = [c for c in schema if str(c["name"]).startswith("column") and str(c["name"])[6:].isdigit()]
    if generic_cols:
        score -= 0.2
        warnings.append(f"{len(generic_cols)} coluna(s) sem cabeçalho (nomes genéricos column0, column1...).")

    # Análise de nulos por coluna
    null_heavy_cols = []
    fully_null_cols = []
    pk_candidate: Optional[str] = None

    for col in schema:
        col_name = col["name"]
        # Quote identifier para nomes com espaço/símbolo. DuckDB usa "".
        qcol = '"' + col_name.replace('"', '""') + '"'
        try:
            r = con.execute(
                f"SELECT COUNT(*) AS total, COUNT({qcol}) AS not_null, "
                f"COUNT(DISTINCT {qcol}) AS distinct_v FROM {_DUCKDB_TABLE}"
            ).fetchone()
            total, not_null, distinct_v = r[0], r[1], r[2]
        except Exception:
            continue

        if total == 0:
            continue
        null_pct = 1.0 - (not_null / total)
        if null_pct >= 0.99:
            fully_null_cols.append(col_name)
        elif null_pct >= 0.5:
            null_heavy_cols.append((col_name, null_pct))

        # PK: primeira coluna 100% única E 100% não-nula
        if pk_candidate is None and not_null == total and distinct_v == total and total > 1:
            pk_candidate = col_name

    if fully_null_cols:
        score -= 0.3
        warnings.append(f"{len(fully_null_cols)} coluna(s) 100% nulas: {', '.join(fully_null_cols[:3])}.")
    if null_heavy_cols:
        score -= 0.2
        sample = [f"{n} ({pct:.0%})" for n, pct in null_heavy_cols[:3]]
        warnings.append(f"{len(null_heavy_cols)} coluna(s) com mais de 50% nulos: {', '.join(sample)}.")

    score = max(0.0, min(1.0, score))
    return round(score, 3), warnings, pk_candidate


def _validate_size(data: bytes) -> None:
    """Rejeita arquivos acima do limite ANTES de carregar."""
    size_mb = len(data) / (1024 * 1024)
    if size_mb > MAX_TABLE_SIZE_MB:
        raise TabularError(
            f"Arquivo {size_mb:.1f}MB excede limite de {MAX_TABLE_SIZE_MB}MB. "
            f"Reduza o arquivo (filtre linhas/colunas) ou peça aumento do limite.",
            status_code=413,
        )


# ─── 1. ANALYZE ──────────────────────────────────────────────────


def _analyze_one_sheet(
    duckdb,
    tmp_path: str,
    ext: str,
    sheet_name: Optional[str],
    header_row: int,
) -> dict:
    """Analisa UMA aba (ou o CSV inteiro). Retorna dict de uma sheet entry.

    NÃO faz auto-retry — só uma tentativa com os params dados. O caller
    decide se quer chamar de novo com header_row diferente.
    """
    con = duckdb.connect(":memory:")
    try:
        _read_into_duckdb(con, tmp_path, ext, sheet_name=sheet_name, header_row=header_row)
        schema = _describe_schema(con)
        if len(schema) > MAX_COLUMNS:
            raise TabularError(
                f"Planilha {('aba ' + sheet_name) if sheet_name else ''} tem {len(schema)} "
                f"colunas, excede limite de {MAX_COLUMNS}. Reduza/divida o arquivo.",
                status_code=413,
            )
        row_count = con.execute(f"SELECT COUNT(*) FROM {_DUCKDB_TABLE}").fetchone()[0]
        score, warnings, pk = _compute_quality_score(con, schema, row_count)
        from app.data_tables.types import TABULAR_READY_THRESHOLD
        return {
            "name": sheet_name or "_csv_",
            "header_row": header_row,
            "rows": row_count,
            "columns": len(schema),
            "schema": schema,
            "score": score,
            "tabular_ready": score >= TABULAR_READY_THRESHOLD,
            "warnings": warnings,
            "suggested_pk": pk,
        }
    finally:
        con.close()


def _analyze_one_sheet_with_retry(
    duckdb,
    tmp_path: str,
    ext: str,
    sheet_name: Optional[str],
) -> dict:
    """Auto-detect do header_row: tenta header_row=1 (padrão). Se resultado
    é patológico (score=0 indicando '1 col VARCHAR única', sinal típico de
    XLSX com título mergeado na linha 1), tenta header_row=2 e mantém o melhor.

    Propaga TabularError e exceções de leitura — `analyze_tabular` decide
    se trata como "arquivo corrupto" (1 aba) ou "1 aba quebrada de N" (multi).
    """
    attempt1 = _analyze_one_sheet(duckdb, tmp_path, ext, sheet_name, header_row=1)

    # Patológico — qualquer um destes sinais dispara auto-retry com header_row=2:
    #   1. Alguma coluna tem nome típico de TABELA (TB_/FATO_/DIM_/...) —
    #      sinal forte e independente de score (cobre o caso real do user
    #      onde só a coluna 0 tem nome de tabela e o score fica em 0.3,
    #      não bate o threshold antigo de score<0.3 mas é claramente patológico).
    #   2. Warning específico de "1 col VARCHAR única" (caso extremo).
    #   3. Score baixo (< threshold) E warning de título mergeado.
    has_title_like_col = any(
        _looks_like_table_title(c["name"]) for c in attempt1.get("schema", [])
    )
    has_single_varchar_warning = any(
        "1 coluna VARCHAR única" in w for w in attempt1["warnings"]
    )
    has_merged_title_warning = any(
        "título mergeado" in w or "linha 1 do XLSX" in w
        for w in attempt1["warnings"]
    )
    from app.data_tables.types import TABULAR_READY_THRESHOLD
    looks_patho = (
        has_title_like_col
        or has_single_varchar_warning
        or (attempt1["score"] < TABULAR_READY_THRESHOLD and has_merged_title_warning)
    )
    if not looks_patho:
        attempt1["header_row_auto_detected"] = False
        return attempt1

    # Retry com header_row=2 (best-effort: se falhar, devolve attempt1)
    try:
        attempt2 = _analyze_one_sheet(duckdb, tmp_path, ext, sheet_name, header_row=2)
    except Exception:
        attempt1["header_row_auto_detected"] = False
        return attempt1

    if attempt2["score"] > attempt1["score"]:
        attempt2["header_row_auto_detected"] = True
        attempt2["warnings"].insert(
            0,
            "Auto-detect: linha 1 parecia título mergeado, header foi recalculado a partir da linha 2."
        )
        return attempt2
    attempt1["header_row_auto_detected"] = False
    return attempt1


async def analyze_tabular(data: bytes, filename: str, mime_type: Optional[str] = None) -> dict:
    """Dry-run: carrega em DuckDB :memory:, infere schema, calcula score.

    Para XLSX multi-aba, analisa CADA aba individualmente com auto-detect
    do header_row. Retorna a "melhor aba" no top-level (compat com legacy)
    + lista completa em `sheets[]`.

    NÃO persiste arquivo nem grava em Postgres.

    Returns:
        {
          "ext": "csv" | "xlsx",
          "size_bytes": int,
          "sheet_count": int,          # 1 para CSV; N para XLSX
          # Atalhos da aba "primária" (maior score):
          "tabular_ready": bool,
          "score": float,
          "rows": int,
          "columns": int,
          "schema": [{name, type, nullable}],
          "warnings": [str, ...],
          "suggested_pk": str | None,
          "primary_sheet": str | None,  # nome da aba primária, None para CSV
          # XLSX: detalhe de cada aba:
          "sheets": [
            {name, header_row, header_row_auto_detected, rows, columns,
             schema, score, tabular_ready, warnings, suggested_pk}
          ],
        }
    """
    _validate_size(data)
    ext = _ext_from_filename(filename)
    if ext == "unsupported":
        raise TabularError(
            f"Extensão não reconhecida em '{filename}'. Aceitos: .csv, .xlsx.",
            status_code=400,
        )

    duckdb = _import_duckdb()
    _t0_analyze = time.perf_counter()
    _tabular_logger.info(
        "analyze_started",
        extra={"event": EVT_ANALYZE_STARTED, "file_name": filename,
               "ext": ext, "size_bytes": len(data)},
    )

    def _run() -> dict:
        with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name
        try:
            if ext == "csv":
                # CSV: 1 só "aba" virtual
                sheet = _analyze_one_sheet_with_retry(duckdb, tmp_path, ext, sheet_name=None)
                return _build_response(ext, len(data), [sheet], primary_sheet=None)

            # XLSX: lista abas e analisa cada uma
            sheets = _list_xlsx_sheets(tmp_path)
            if not sheets:
                # Fallback: openpyxl não conseguiu → trata como aba única default
                sheet = _analyze_one_sheet_with_retry(duckdb, tmp_path, ext, sheet_name=None)
                return _build_response(ext, len(data), [sheet], primary_sheet=None)

            analyses = []
            for name in sheets:
                try:
                    s = _analyze_one_sheet_with_retry(duckdb, tmp_path, ext, sheet_name=name)
                except TabularError as e:
                    # 1 aba quebrada não derruba a análise inteira
                    s = {
                        "name": name, "header_row": 1, "rows": 0, "columns": 0,
                        "schema": [], "score": 0.0, "tabular_ready": False,
                        "warnings": [str(e)], "suggested_pk": None,
                        "header_row_auto_detected": False, "error": True,
                    }
                analyses.append(s)

            # Aba primária = maior score (desempate: mais colunas)
            primary = max(analyses, key=lambda s: (s["score"], s["columns"]))
            return _build_response(ext, len(data), analyses, primary_sheet=primary["name"])
        except TabularError:
            raise
        except Exception as e:
            raise TabularError(
                f"Falha ao analisar arquivo: {e}. "
                f"Verifique se o CSV/XLSX está bem-formado.",
                status_code=400,
            ) from e
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    try:
        result = await asyncio.to_thread(_run)
    except TabularError as e:
        _tabular_logger.warning(
            "analyze_failed",
            extra={
                "event": EVT_ANALYZE_FAILED, "file_name": filename, "ext": ext,
                "error_class": type(e).__name__, "error_msg": str(e),
                "status_code": e.status_code,
                "duration_ms": round((time.perf_counter() - _t0_analyze) * 1000, 2),
            },
        )
        raise

    _tabular_logger.info(
        "analyze_completed",
        extra={
            "event": EVT_ANALYZE_COMPLETED, "file_name": filename, "ext": ext,
            "sheet_count": result.get("sheet_count"),
            "primary_sheet": result.get("primary_sheet"),
            "top_score": result.get("score"),
            "any_ready": any(s.get("tabular_ready") for s in result.get("sheets", [])),
            "has_auto_detect": any(s.get("header_row_auto_detected")
                                   for s in result.get("sheets", [])),
            "duration_ms": round((time.perf_counter() - _t0_analyze) * 1000, 2),
        },
    )
    return result


def _build_response(
    ext: str,
    size_bytes: int,
    sheets: list[dict],
    primary_sheet: Optional[str],
) -> dict:
    """Monta dict de resposta a partir das análises de cada aba.

    Top-level reflete a "aba primária" (maior score, com desempate por colunas).
    Para CSV, há sempre 1 aba virtual e `primary_sheet=None`.
    """
    if not sheets:
        return {
            "ext": ext, "size_bytes": size_bytes, "sheet_count": 0,
            "tabular_ready": False, "score": 0.0, "rows": 0, "columns": 0,
            "schema": [], "warnings": ["Nenhuma aba legível encontrada."],
            "suggested_pk": None, "primary_sheet": None, "sheets": [],
        }
    if primary_sheet is None:
        primary = sheets[0]
    else:
        primary = next((s for s in sheets if s["name"] == primary_sheet), sheets[0])
    return {
        "ext": ext,
        "size_bytes": size_bytes,
        "sheet_count": len(sheets),
        "tabular_ready": primary["tabular_ready"],
        "score": primary["score"],
        "rows": primary["rows"],
        "columns": primary["columns"],
        "schema": primary["schema"],
        "warnings": primary["warnings"],
        "suggested_pk": primary["suggested_pk"],
        "primary_sheet": primary_sheet,
        "sheets": sheets,
    }


# ─── 2. PROMOTE ──────────────────────────────────────────────────


async def promote_to_table(
    ks_id: str,
    data: bytes,
    filename: str,
    name: Optional[str] = None,
    description: str = "",
    created_by: Optional[str] = None,
    sheet_name: Optional[str] = None,
    header_row: Optional[int] = None,
) -> dict:
    """Cria arquivo .duckdb persistente e registra em data_tables.

    Args:
        ks_id: knowledge_source de origem. Deve existir.
        data: bytes do CSV/XLSX.
        filename: nome original (usado para slug + extensão).
        name: nome amigável da tabela. Default = filename sem extensão.
        description: descrição livre.
        created_by: user id (para audit).
        sheet_name: nome da aba do XLSX a promover. None = aba primária
            (a com maior score na análise). Para promover N abas, faça
            N chamadas com sheet_name diferente em cada.
        header_row: linha (1-based) com os headers. None = usar o que a
            análise auto-detectou para essa aba. Forçar útil se o user
            quiser override manual.

    Returns:
        Dict da tabela criada (com `id`, `urn`, `schema_json`, etc.)

    Raises:
        TabularError: ks não existe (404), arquivo inválido (400),
                      tamanho excedido (413), DuckDB falha (500).
    """
    _t0_promote = time.perf_counter()
    _tabular_logger.info(
        "promote_started",
        extra={"event": EVT_PROMOTE_STARTED, "ks_id": ks_id, "file_name": filename,
               "sheet_name": sheet_name, "header_row": header_row},
    )
    # Sanity checks
    ks = await knowledge_repo.find_by_id(ks_id)
    if not ks:
        _tabular_logger.warning(
            "promote_failed",
            extra={"event": EVT_PROMOTE_FAILED, "ks_id": ks_id,
                   "file_name": filename, "error_class": "TabularError",
                   "error_msg": "knowledge_source not found", "status_code": 404,
                   "duration_ms": round((time.perf_counter() - _t0_promote) * 1000, 2)},
        )
        raise TabularError(f"knowledge_source '{ks_id}' não encontrada.", status_code=404)

    # Reusa análise para validar + capturar schema/score por aba
    analysis = await analyze_tabular(data, filename)
    ext = analysis["ext"]

    # Resolve qual aba + header_row efetivos:
    # - se user pediu sheet_name explícito, busca essa aba na análise
    # - senão usa a "primária" (top-level já reflete ela)
    target_sheet: dict
    if sheet_name and ext == "xlsx":
        match = next((s for s in analysis.get("sheets", []) if s["name"] == sheet_name), None)
        if not match:
            available = [s["name"] for s in analysis.get("sheets", [])]
            raise TabularError(
                f"Aba '{sheet_name}' não encontrada. Disponíveis: {available}",
                status_code=400,
            )
        target_sheet = match
    else:
        # CSV ou XLSX sem sheet_name: usa a primária
        if ext == "xlsx" and analysis.get("sheets"):
            primary_name = analysis.get("primary_sheet")
            target_sheet = next(
                (s for s in analysis["sheets"] if s["name"] == primary_name),
                analysis["sheets"][0],
            )
            sheet_name = target_sheet["name"]
        else:
            target_sheet = {
                "schema": analysis["schema"],
                "rows": analysis["rows"],
                "columns": analysis["columns"],
                "header_row": 1,
                "suggested_pk": analysis.get("suggested_pk"),
                "score": analysis["score"],
            }

    effective_header_row = header_row if header_row is not None else target_sheet.get("header_row", 1)
    schema = target_sheet["schema"]

    # Nome amigável: user explícito > "filename — sheet" (multi-aba) > filename
    base_filename = os.path.splitext(os.path.basename(filename))[0] or "tabela"
    is_multi_sheet = ext == "xlsx" and analysis.get("sheet_count", 1) > 1
    if name:
        display_name = name
    elif is_multi_sheet and sheet_name:
        display_name = f"{base_filename} — {sheet_name}"
    else:
        display_name = base_filename

    # Slug + versão. Para XLSX multi-aba, sufixa com sheet_name no slug
    # para diferenciar a URN entre as N abas do mesmo arquivo.
    if is_multi_sheet and sheet_name:
        slug = slugify(f"{name or base_filename}__{sheet_name}")
    else:
        slug = slugify(name or base_filename)
    version = await next_version_for_slug(ks_id, slug)
    urn = build_urn(ks_id, slug, version)

    # Caminhos
    table_id = str(uuid.uuid4())
    ks_dir = _TABULAR_ROOT / ks_id
    duckdb_path = ks_dir / f"{table_id}.duckdb"
    relative_path = str(duckdb_path).replace("\\", "/")

    duckdb = _import_duckdb()

    def _write() -> int:
        """Cria o .duckdb e ingere os dados. Retorna size_bytes do arquivo final."""
        ks_dir.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name
        try:
            con = duckdb.connect(str(duckdb_path))
            try:
                _read_into_duckdb(
                    con, tmp_path, ext,
                    sheet_name=sheet_name if ext == "xlsx" else None,
                    header_row=effective_header_row,
                )
            finally:
                con.close()
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
        return duckdb_path.stat().st_size

    try:
        size_bytes = await asyncio.to_thread(_write)
    except TabularError:
        raise
    except Exception as e:
        # Cleanup parcial
        try:
            duckdb_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise TabularError(f"Falha ao gravar DuckDB: {e}", status_code=500) from e

    # Persiste metadata em Postgres. Usa dados da target_sheet (não top-level
    # da análise) — se o user escolheu uma aba diferente da primária, esses
    # valores divergem.
    row = {
        "id": table_id,
        "knowledge_source_id": ks_id,
        "urn": urn,
        "name": display_name,
        "description": description,
        # JSONB: asyncpg exige string JSON (não list/dict Python) ao passar via
        # placeholder $N. Repository genérico não faz encoding — fazemos aqui.
        "schema_json": json.dumps(schema, ensure_ascii=False),
        "row_count": target_sheet.get("rows", 0),
        "column_count": target_sheet.get("columns", len(schema)),
        "size_bytes": size_bytes,
        "duckdb_path": relative_path,
        "duckdb_table_name": _DUCKDB_TABLE,
        "version": str(version),
        "status": "ready",
        "quality_score": target_sheet.get("score", 0.0),
        "suggested_pk": target_sheet.get("suggested_pk"),
        "created_by": created_by or "",
    }
    try:
        await data_tables_repo.create(row)
    except Exception as e:
        # Rollback: apaga arquivo se metadata falhar (consistência)
        try:
            duckdb_path.unlink(missing_ok=True)
        except OSError:
            pass
        _tabular_logger.error(
            "promote_failed",
            extra={"event": EVT_PROMOTE_FAILED, "ks_id": ks_id, "file_name": filename,
                   "sheet_name": sheet_name, "error_class": type(e).__name__,
                   "error_msg": str(e)[:200], "status_code": 500,
                   "duration_ms": round((time.perf_counter() - _t0_promote) * 1000, 2)},
        )
        raise TabularError(f"Falha ao registrar tabela: {e}", status_code=500) from e

    # Retorna versão enriquecida (com flags da KS) para a UI
    enriched = await find_by_id_with_ks(table_id)

    # Log de sucesso (catálogo de eventos)
    # IMPORTANTE: as chaves do `extra` NÃO podem colidir com atributos
    # reservados de LogRecord (name, msg, args, levelname, message, module,
    # funcName, lineno, etc) — Python levanta KeyError em makeRecord.
    # Por isso `table_name` em vez de `name` (PR #225 corrigiu o bug original).
    _tabular_logger.info(
        "promote_completed",
        extra={
            "event": EVT_PROMOTE_COMPLETED,
            "ks_id": ks_id, "table_id": table_id, "urn": urn,
            "table_name": display_name, "sheet_name": sheet_name,
            "rows": target_sheet.get("rows", 0),
            "columns": target_sheet.get("columns", len(schema)),
            "size_bytes": size_bytes,
            "quality_score": target_sheet.get("score", 0.0),
            "suggested_pk": target_sheet.get("suggested_pk"),
            "duration_ms": round((time.perf_counter() - _t0_promote) * 1000, 2),
        },
    )
    return enriched or row


# ─── 3. EXECUTE QUERY ────────────────────────────────────────────


def _quote_ident(name: str) -> str:
    """Escapa identificador (col/table) para DuckDB usando double-quote."""
    return '"' + (name or "").replace('"', '""') + '"'


def _validate_columns(requested: list[str], schema: list[dict]) -> None:
    """Garante que toda coluna solicitada existe no schema. Erro 400 se não."""
    valid = {c["name"] for c in schema}
    unknown = [c for c in requested if c not in valid]
    if unknown:
        raise TabularError(
            f"Coluna(s) inexistente(s): {', '.join(unknown)}. "
            f"Disponíveis: {', '.join(sorted(valid))}.",
            status_code=400,
        )


# ─── 2b. APPEND (incremento em tabela existente) ─────────────────


async def append_to_table(
    target_table_id: str,
    data: bytes,
    filename: str,
    sheet_name: Optional[str] = None,
    header_row: Optional[int] = None,
) -> dict:
    """Adiciona linhas a uma data_table EXISTENTE (incremento).

    Diferente de `promote_to_table` (que cria nova versão), esta função
    abre o arquivo .duckdb da tabela em modo WRITE e faz INSERT de novas
    linhas vindas de outro CSV/XLSX. Schema NÃO muda — colunas extras no
    arquivo novo são ignoradas; colunas faltantes ficam NULL.

    Args:
        target_table_id: id da data_table que receberá os dados novos.
        data: bytes do CSV/XLSX com novas linhas.
        filename: nome original (para detectar ext).
        sheet_name: aba do XLSX (None = primeira). Ignorado pra CSV.
        header_row: 1-based, linha do header. None = auto-detect padrão
            (header_row=1, sem retry — espera-se que tenha mesma estrutura
            da tabela existente).

    Returns:
        {
          "table_id": <id>,
          "rows_added": N,
          "row_count_before": N,
          "row_count_after": N,
          "duration_ms": int,
        }

    Raises:
        TabularError: tabela não existe (404), arquivo inválido (400),
                      schema incompatível (400), DuckDB write falha (500).
    """
    target = await find_by_id_with_ks(target_table_id)
    if not target:
        raise TabularError(
            f"data_table '{target_table_id}' não encontrada.",
            status_code=404,
        )
    if target.get("status") != "ready":
        raise TabularError(
            f"Tabela '{target_table_id}' não está pronta (status={target.get('status')}).",
            status_code=409,
        )

    _validate_size(data)
    ext = _ext_from_filename(filename)
    if ext == "unsupported":
        raise TabularError(
            f"Extensão não reconhecida em '{filename}'. Aceitos: .csv, .xlsx.",
            status_code=400,
        )

    duckdb_path = target.get("duckdb_path")
    if not duckdb_path or not Path(duckdb_path).exists():
        raise TabularError(
            f"Arquivo DuckDB ausente para tabela '{target_table_id}': {duckdb_path}",
            status_code=500,
        )

    # Schema da tabela target — colunas a preservar
    raw_schema = target.get("schema_json") or []
    if isinstance(raw_schema, str):
        try:
            target_schema = json.loads(raw_schema)
        except (json.JSONDecodeError, TypeError):
            target_schema = []
    else:
        target_schema = raw_schema
    target_cols = [c["name"] for c in target_schema]
    if not target_cols:
        raise TabularError(
            f"Schema da tabela '{target_table_id}' está vazio — não dá pra appendar.",
            status_code=500,
        )

    duckdb = _import_duckdb()
    eff_header_row = header_row if header_row is not None else 1

    # Estratégia: criar staging TEMPORARY (não conflita com 'data'), copiar
    # SOMENTE as colunas em comum (intersecção de schema) para 'data'.
    # Colunas extras no arquivo novo são ignoradas; faltantes ficam NULL.
    def _do_append() -> tuple[int, int]:
        with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name
        try:
            # Para XLSX: converter para CSV primeiro (pipeline padrão)
            csv_to_import: str
            if ext == "xlsx":
                csv_handle = tempfile.NamedTemporaryFile(
                    suffix=".csv", delete=False, mode="w",
                    encoding="utf-8", newline="",
                )
                csv_to_import = csv_handle.name
                csv_handle.close()
                _xlsx_sheet_to_csv(tmp_path, sheet_name, eff_header_row, csv_to_import)
                skip_clause = ""  # CSV gerado já tem header na linha 1
            else:
                csv_to_import = tmp_path
                skip = max(0, eff_header_row - 1)
                skip_clause = f", SKIP={skip}" if skip > 0 else ""

            con = duckdb.connect(duckdb_path)
            try:
                count_before = con.execute(
                    f"SELECT COUNT(*) FROM {_DUCKDB_TABLE}"
                ).fetchone()[0]

                safe = csv_to_import.replace("'", "''")
                # Cria staging table com nome único
                con.execute(
                    f"CREATE OR REPLACE TEMPORARY TABLE _staging AS "
                    f"SELECT * FROM read_csv_auto('{safe}', HEADER=TRUE{skip_clause})"
                )
                # Pega colunas que existem em AMBAS (intersecção)
                staging_cols = [r[0] for r in con.execute(
                    "DESCRIBE _staging"
                ).fetchall()]
                common = [c for c in target_cols if c in staging_cols]
                if not common:
                    raise TabularError(
                        f"Schema incompatível: nenhuma coluna em comum. "
                        f"Tabela espera: {target_cols}. Arquivo tem: {staging_cols}.",
                        status_code=400,
                    )
                quoted = ", ".join(_quote_ident(c) for c in common)
                con.execute(
                    f"INSERT INTO {_DUCKDB_TABLE} ({quoted}) "
                    f"SELECT {quoted} FROM _staging"
                )
                count_after = con.execute(
                    f"SELECT COUNT(*) FROM {_DUCKDB_TABLE}"
                ).fetchone()[0]
                con.execute("DROP TABLE IF EXISTS _staging")
                return count_before, count_after
            finally:
                con.close()
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            if ext == "xlsx" and csv_to_import != tmp_path:
                try:
                    os.unlink(csv_to_import)
                except OSError:
                    pass

    _tabular_logger.info(
        "append_started",
        extra={"event": EVT_APPEND_STARTED, "table_id": target_table_id,
               "file_name": filename, "sheet_name": sheet_name},
    )
    t0 = time.perf_counter()
    try:
        count_before, count_after = await asyncio.to_thread(_do_append)
    except TabularError as e:
        _tabular_logger.warning(
            "append_failed",
            extra={"event": EVT_APPEND_FAILED, "table_id": target_table_id,
                   "file_name": filename, "error_class": type(e).__name__,
                   "error_msg": str(e)[:200], "status_code": e.status_code,
                   "duration_ms": int((time.perf_counter() - t0) * 1000)},
        )
        raise
    except Exception as e:
        _tabular_logger.error(
            "append_failed",
            extra={"event": EVT_APPEND_FAILED, "table_id": target_table_id,
                   "file_name": filename, "error_class": type(e).__name__,
                   "error_msg": str(e)[:200], "status_code": 500,
                   "duration_ms": int((time.perf_counter() - t0) * 1000)},
        )
        raise TabularError(f"Falha ao appendar: {e}", status_code=500) from e
    duration_ms = int((time.perf_counter() - t0) * 1000)
    rows_added = count_after - count_before
    _tabular_logger.info(
        "append_completed",
        extra={"event": EVT_APPEND_COMPLETED, "table_id": target_table_id,
               "rows_added": rows_added, "row_count_before": count_before,
               "row_count_after": count_after, "duration_ms": duration_ms},
    )

    # Atualiza row_count e size_bytes na metadata
    try:
        new_size = Path(duckdb_path).stat().st_size
        await data_tables_repo.update(target_table_id, {
            "row_count": count_after,
            "size_bytes": new_size,
            "updated_at": __import__("datetime").datetime.utcnow(),
        })
    except Exception as e:
        logger.warning("append_to_table: failed to update metadata: %s", e)

    return {
        "table_id": target_table_id,
        "rows_added": rows_added,
        "row_count_before": count_before,
        "row_count_after": count_after,
        "duration_ms": duration_ms,
    }


async def execute_query(
    table_id: str,
    inputs: Optional[dict] = None,
    select: Optional[list[str]] = None,
    filters: Optional[list[dict]] = None,
    order_by: Optional[list[str]] = None,
    limit: int = 100,
    executed_by: Optional[str] = None,
    interaction_id: Optional[str] = None,
    agent_id: str = "",
) -> dict:
    """Executa SELECT parametrizado em uma data_table.

    Args:
        table_id: id da data_table.
        inputs: dict de inputs do usuário (paridade com declarative_engine).
                Filtros com `if_present: X` são pulados se X não estiver em inputs.
        select: colunas a retornar. None = todas.
        filters: list[{col, op, value, if_present?}]. `op` é string do SqlOperator.
                 `value` pode ser literal OU template "{{ inputs.X }}" (resolvido aqui).
        order_by: list de "col" ou "col DESC". Validado contra schema.
        limit: máximo de linhas. Hard-cap = MAX_ROWS_RETURNED.
        executed_by, interaction_id, agent_id: contexto pra auditoria.

    Returns:
        {
          "rows": [{col: val, ...}, ...],
          "row_count": N,
          "columns": [col_names],
          "duration_ms": int,
          "sql_rendered": str,        # SQL com ? (sem bind values)
          "table": {id, urn, name},
        }

    Raises:
        TabularError: tabela não existe (404), coluna inválida (400),
                      execução falha (500), limit > cap (400).
    """
    inputs = inputs or {}
    filters = filters or []
    select = select or []
    order_by = order_by or []

    table = await find_by_id_with_ks(table_id)
    if not table:
        raise TabularError(f"data_table '{table_id}' não encontrada.", status_code=404)
    if table.get("status") != "ready":
        raise TabularError(
            f"Tabela '{table_id}' não está pronta (status={table.get('status')}).",
            status_code=409,
        )

    if limit > MAX_ROWS_RETURNED:
        raise TabularError(
            f"limit={limit} excede máximo {MAX_ROWS_RETURNED}.",
            status_code=400,
        )
    if limit < 1:
        limit = 1

    # Em produção, asyncpg decoda JSONB → list nativo. Em testes (mocks que
    # guardam o que foi passado pra create), pode vir como string JSON.
    # Decode defensivo cobre os 2 caminhos.
    raw_schema = table.get("schema_json") or []
    if isinstance(raw_schema, str):
        try:
            schema = json.loads(raw_schema)
        except (json.JSONDecodeError, TypeError):
            schema = []
    else:
        schema = raw_schema
    schema_col_names = [c["name"] for c in schema]

    # SELECT clause
    if select:
        _validate_columns(select, schema)
        select_sql = ", ".join(_quote_ident(c) for c in select)
        result_columns = list(select)
    else:
        select_sql = "*"
        result_columns = list(schema_col_names)

    # WHERE clause
    where_parts: list[str] = []
    bind_values: list[Any] = []

    for f in filters:
        col = f.get("col") or f.get("column")
        op_raw = f.get("op") or f.get("operator")
        if not col or not op_raw:
            raise TabularError(f"Filtro inválido (col + op obrigatórios): {f}", status_code=400)

        # if_present: skip filter se input não fornecido
        if_present_key = f.get("if_present")
        if if_present_key and inputs.get(if_present_key) in (None, ""):
            continue

        # Validar operador (defense in depth)
        try:
            op = SqlOperator(op_raw if isinstance(op_raw, str) else op_raw.value)
        except ValueError:
            raise TabularError(
                f"Operador '{op_raw}' não suportado. Aceitos: {[o.value for o in SqlOperator]}.",
                status_code=400,
            )

        # Validar coluna
        _validate_columns([col], schema)

        # Resolver value (suporta template simples "{{ inputs.X }}")
        raw_value = f.get("value")
        value = _resolve_template_value(raw_value, inputs)

        clause, binds = render_where_clause(_quote_ident(col), op, value)
        where_parts.append(clause)
        bind_values.extend(binds)

    where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""

    # ORDER BY
    order_sql = ""
    if order_by:
        ord_parts = []
        for ob in order_by:
            tokens = str(ob).strip().split()
            ob_col = tokens[0]
            direction = tokens[1].upper() if len(tokens) > 1 else "ASC"
            if direction not in ("ASC", "DESC"):
                raise TabularError(f"Direção inválida em ORDER BY: {ob}", status_code=400)
            _validate_columns([ob_col], schema)
            ord_parts.append(f"{_quote_ident(ob_col)} {direction}")
        order_sql = " ORDER BY " + ", ".join(ord_parts)

    sql = (
        f"SELECT {select_sql} FROM {_DUCKDB_TABLE}"
        f"{where_sql}{order_sql} LIMIT ?"
    )
    bind_values_final = list(bind_values) + [limit]

    duckdb = _import_duckdb()
    duckdb_path = table.get("duckdb_path")
    if not duckdb_path:
        raise TabularError("duckdb_path ausente na tabela.", status_code=500)

    def _run_query() -> tuple[list[dict], int, list[str]]:
        # READ-ONLY: safety técnica. INSERT/UPDATE/DELETE/DROP rejeitados pelo engine.
        con = duckdb.connect(duckdb_path, read_only=True)
        try:
            cur = con.execute(sql, bind_values_final)
            rows = cur.fetchall()
            col_descr = cur.description or []
            col_names = [d[0] for d in col_descr]
            row_dicts = [dict(zip(col_names, r)) for r in rows]
            return row_dicts, len(row_dicts), col_names
        finally:
            con.close()

    t0 = time.perf_counter()
    status = "ok"
    error_msg: Optional[str] = None
    rows: list[dict] = []
    row_count = 0
    columns_out: list[str] = result_columns

    try:
        rows, row_count, columns_out = await asyncio.to_thread(_run_query)
    except TabularError:
        status = "error"
        raise
    except Exception as e:
        status = "error"
        error_msg = str(e)
        raise TabularError(f"Falha ao executar query: {e}", status_code=500) from e
    finally:
        duration_ms = int((time.perf_counter() - t0) * 1000)
        # Audit (best-effort: falha não derruba a request)
        try:
            await data_table_query_logs_repo.create({
                "id": str(uuid.uuid4()),
                "data_table_id": table_id,
                "interaction_id": interaction_id,
                "agent_id": agent_id,
                "executed_by": executed_by or "",
                "sql_rendered": sql,
                # JSONB: encoding explícito (mesma razão do schema_json em promote)
                "inputs_json": json.dumps(inputs, ensure_ascii=False, default=str),
                "row_count": row_count,
                "duration_ms": duration_ms,
                "status": status,
                "error_message": error_msg,
            })
        except Exception:
            logger.exception("Falha ao gravar audit log de query tabular")

    # Evento canônico: tabular.query.executed (success ou error)
    operators_used = sorted({
        str(f.get("op", "")) for f in (filters or []) if f.get("op")
    })
    has_template = any(
        isinstance(f.get("value"), str) and "{{" in f.get("value", "")
        for f in (filters or [])
    )
    _tabular_logger.info(
        "query_executed",
        extra={
            "event": EVT_QUERY_EXECUTED,
            "table_id": table_id,
            "table_urn": table.get("urn"),
            "operators_used": operators_used,
            "select_count": len(select or []),
            "has_template": has_template,
            "row_count": row_count,
            "duration_ms": duration_ms,
            "status": status,
        },
    )

    return {
        "rows": rows,
        "row_count": row_count,
        "columns": columns_out,
        "duration_ms": duration_ms,
        "sql_rendered": sql,
        "table": {
            "id": table.get("id"),
            "urn": table.get("urn"),
            "name": table.get("name"),
        },
    }


_TEMPLATE_RE = None  # lazy compile


def _resolve_template_value(value: Any, inputs: dict) -> Any:
    """Resolve template MUITO simples: "{{ inputs.foo }}" → inputs["foo"].

    Não usa Jinja2 para evitar overhead e superfície de injeção. O engine
    declarativo (declarative_engine) já usa Jinja2 com SandboxedEnvironment
    no caminho de produção; este resolver é apenas para o caminho de teste
    direto via UI Query Builder.
    """
    if not isinstance(value, str):
        return value
    s = value.strip()
    if s.startswith("{{") and s.endswith("}}"):
        expr = s[2:-2].strip()
        if expr.startswith("inputs."):
            key = expr[len("inputs."):].strip()
            return inputs.get(key)
    return value
