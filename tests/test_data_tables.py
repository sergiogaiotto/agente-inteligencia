"""Testes da Onda Tabular — analyze, promote, query, routes.

Estratégia:
- Helpers puros (operators, slugify, visibility): tests sem mock.
- analyze_tabular / execute_query: DuckDB real em :memory: ou tmp file.
- promote_to_table: monkeypatch dos repos Postgres (sem subir banco).
- Routes: mini FastAPI + dependency_overrides[require_user] + monkeypatch
  das funções que tocam Postgres (find_by_id_with_ks, list_for_user, etc.).

DuckDB é pré-requisito (em requirements.txt). Se não estiver instalado,
todos os testes que importam o service vão skip via pytest.importorskip.
"""

from __future__ import annotations

import asyncio
import os
import uuid
from pathlib import Path
from typing import Any, Optional

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

# Skip toda a suite se DuckDB indisponível (não derruba CI)
duckdb = pytest.importorskip("duckdb", reason="DuckDB não instalado")

from app.core.auth import require_user
from app.data_tables.queries import (
    build_urn,
    can_user_see,
    is_root,
    slugify,
)
from app.data_tables.types import (
    MAX_ROWS_RETURNED,
    SqlOperator,
    render_where_clause,
)
from app.evidence import tabular as tabular_service
from app.evidence.tabular import (
    TabularError,
    analyze_tabular,
    execute_query,
    promote_to_table,
)
from app.routes.data_tables import router as data_tables_router


# ═══════════════════════════════════════════════════════════════
# 1. Pure helpers — types.py
# ═══════════════════════════════════════════════════════════════


class TestSqlOperatorRendering:
    def test_eq_renders_with_one_bind(self):
        clause, binds = render_where_clause("nome", SqlOperator.EQ, "alice")
        assert clause == "nome = ?"
        assert binds == ["alice"]

    def test_neq_renders(self):
        clause, binds = render_where_clause("col", SqlOperator.NEQ, 42)
        assert clause == "col != ?"
        assert binds == [42]

    def test_gte_renders(self):
        clause, binds = render_where_clause("idade", SqlOperator.GTE, 18)
        assert clause == "idade >= ?"
        assert binds == [18]

    def test_like_renders(self):
        clause, binds = render_where_clause("nome", SqlOperator.LIKE, "Al%")
        assert clause == "nome LIKE ?"
        assert binds == ["Al%"]

    def test_ilike_renders(self):
        clause, binds = render_where_clause("nome", SqlOperator.ILIKE, "al%")
        assert clause == "nome ILIKE ?"
        assert binds == ["al%"]

    def test_in_renders_with_n_placeholders(self):
        clause, binds = render_where_clause("status", SqlOperator.IN, ["a", "b", "c"])
        assert clause == "status IN (?, ?, ?)"
        assert binds == ["a", "b", "c"]

    def test_not_in_renders(self):
        clause, binds = render_where_clause("tipo", SqlOperator.NOT_IN, [1, 2])
        assert clause == "tipo NOT IN (?, ?)"
        assert binds == [1, 2]

    def test_between_renders_with_two_binds(self):
        clause, binds = render_where_clause("idade", SqlOperator.BETWEEN, [18, 65])
        assert clause == "idade BETWEEN ? AND ?"
        assert binds == [18, 65]

    def test_is_null_renders_without_bind(self):
        clause, binds = render_where_clause("email", SqlOperator.IS_NULL, None)
        assert clause == "email IS NULL"
        assert binds == []

    def test_is_not_null_renders_without_bind(self):
        clause, binds = render_where_clause("phone", SqlOperator.IS_NOT_NULL, "ignored")
        assert clause == "phone IS NOT NULL"
        assert binds == []

    def test_between_rejects_non_list_value(self):
        with pytest.raises(ValueError, match="BETWEEN"):
            render_where_clause("col", SqlOperator.BETWEEN, "naive")

    def test_in_rejects_empty_list(self):
        with pytest.raises(ValueError, match="IN"):
            render_where_clause("col", SqlOperator.IN, [])


# ═══════════════════════════════════════════════════════════════
# 2. Pure helpers — queries.py (slugify, urn, visibility)
# ═══════════════════════════════════════════════════════════════


class TestSlugify:
    def test_simple_name(self):
        assert slugify("Vendas") == "vendas"

    def test_strips_accents(self):
        assert slugify("Relatório de Vendas Q4") == "relatorio-de-vendas-q4"

    def test_handles_extension(self):
        assert slugify("dados.csv") == "dados-csv"

    def test_empty_returns_default(self):
        assert slugify("") == "tabela"
        assert slugify(None) == "tabela"  # type: ignore

    def test_only_symbols_returns_default(self):
        assert slugify("@@@!!!") == "tabela"


class TestBuildUrn:
    def test_canonical_format(self):
        urn = build_urn("abcd1234efgh", "vendas-q4", 1)
        assert urn == "urn:table:abcd1234:vendas-q4:1"

    def test_short_ks_id_padded(self):
        urn = build_urn("xyz", "tbl", 2)
        assert urn == "urn:table:xyz:tbl:2"


class TestVisibility:
    def test_root_sees_everything(self):
        user = {"role": "root"}
        table = {"ks_authorized": 0, "ks_confidentiality_label": "confidential"}
        assert can_user_see(user, table) is True

    def test_non_root_blocked_when_ks_unauthorized(self):
        user = {"role": "comum"}
        table = {"ks_authorized": 0, "ks_confidentiality_label": "public"}
        assert can_user_see(user, table) is False

    def test_non_root_sees_public(self):
        user = {"role": "comum"}
        table = {"ks_authorized": 1, "ks_confidentiality_label": "public"}
        assert can_user_see(user, table) is True

    def test_non_root_sees_internal(self):
        user = {"role": "comum"}
        table = {"ks_authorized": 1, "ks_confidentiality_label": "internal"}
        assert can_user_see(user, table) is True

    def test_non_root_blocked_on_restricted(self):
        user = {"role": "comum"}
        table = {"ks_authorized": 1, "ks_confidentiality_label": "restricted"}
        assert can_user_see(user, table) is False

    def test_non_root_blocked_on_confidential(self):
        user = {"role": "comum"}
        table = {"ks_authorized": 1, "ks_confidentiality_label": "confidential"}
        assert can_user_see(user, table) is False

    def test_is_root_case_insensitive(self):
        assert is_root({"role": "Root"}) is True
        assert is_root({"role": "ROOT"}) is True
        assert is_root({"role": "comum"}) is False


# ═══════════════════════════════════════════════════════════════
# 3. analyze_tabular — DuckDB real, sem Postgres
# ═══════════════════════════════════════════════════════════════


CSV_OK = (
    b"id,nome,idade,salario\n"
    b"1,Alice,30,5000.50\n"
    b"2,Bob,25,3200.00\n"
    b"3,Carol,42,8100.75\n"
    b"4,Dan,33,4500.00\n"
)

CSV_WITH_NULLS = (
    b"id,nome,email\n"
    b"1,Alice,\n"
    b"2,Bob,\n"
    b"3,Carol,carol@x.com\n"
)


class TestAnalyzeTabular:
    def test_csv_well_formed_returns_ready(self):
        result = asyncio.run(analyze_tabular(CSV_OK, "test.csv"))
        assert result["tabular_ready"] is True
        assert result["score"] >= 0.7
        assert result["ext"] == "csv"
        assert result["rows"] == 4
        assert result["columns"] == 4
        # Schema deve ter os 4 nomes
        names = [c["name"] for c in result["schema"]]
        assert set(names) == {"id", "nome", "idade", "salario"}

    def test_csv_with_unique_first_col_suggests_pk(self):
        result = asyncio.run(analyze_tabular(CSV_OK, "test.csv"))
        assert result["suggested_pk"] == "id"

    def test_csv_with_many_nulls_warns(self):
        result = asyncio.run(analyze_tabular(CSV_WITH_NULLS, "nulls.csv"))
        # Coluna email tem 2/3 nulos → > 50% → warning
        assert any("nulos" in w.lower() or "50%" in w for w in result["warnings"])

    def test_binary_file_rejected(self):
        # Bytes que não parecem CSV — DuckDB read_csv_auto vai falhar.
        # Extension .csv passa filtro inicial, mas o parse vai estourar.
        garbage = b"\x00\x01\x02\xff\xfe" * 100
        with pytest.raises(TabularError) as exc:
            asyncio.run(analyze_tabular(garbage, "fake.csv"))
        # Pode falhar com 400 (bad CSV) ou 413 (size) — ambos OK
        assert exc.value.status_code in (400, 413)

    def test_unsupported_extension_rejected(self):
        with pytest.raises(TabularError) as exc:
            asyncio.run(analyze_tabular(b"any data", "doc.pdf"))
        assert exc.value.status_code == 400
        assert "extens" in str(exc.value).lower()

    def test_oversized_file_rejected(self, monkeypatch):
        # Patch MAX_TABLE_SIZE_MB para 0.001 → 1KB
        monkeypatch.setattr(tabular_service, "MAX_TABLE_SIZE_MB", 0.001)
        big = b"col1,col2\n" + (b"x,y\n" * 1000)  # > 1KB
        with pytest.raises(TabularError) as exc:
            asyncio.run(analyze_tabular(big, "big.csv"))
        assert exc.value.status_code == 413


# ═══════════════════════════════════════════════════════════════
# 3b. analyze_tabular — XLSX multi-aba + header mergeado
# ═══════════════════════════════════════════════════════════════


def _build_patho_xlsx(path):
    """Cria XLSX com 2 abas, cada uma com título mergeado na linha 1
    e headers reais na linha 2. Simula o caso real reportado pelo user.
    """
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "TB_ANALISE_CREDITO"
    ws1.merge_cells("A1:D1")
    ws1["A1"] = "TB_ANALISE_CREDITO"
    ws1["A1"].font = Font(bold=True)
    headers1 = ["cd_cliente", "nr_idade", "vr_renda", "vr_limite"]
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=2, column=i, value=h)
    for r in range(3, 13):
        ws1.cell(row=r, column=1, value=r - 2)
        ws1.cell(row=r, column=2, value=20 + r)
        ws1.cell(row=r, column=3, value=1000.0 * r)
        ws1.cell(row=r, column=4, value=5000 + r * 100)

    ws2 = wb.create_sheet("TB_PREDICOES")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "TB_PREDICOES"
    headers2 = ["cd_cliente", "score", "decisao"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i, value=h)
    for r in range(3, 13):
        ws2.cell(row=r, column=1, value=r - 2)
        ws2.cell(row=r, column=2, value=0.1 * r)
        ws2.cell(row=r, column=3, value="aprovado" if r % 2 == 0 else "rejeitado")
    wb.save(path)


class TestAnalyzeXlsxMultiSheet:
    def test_xlsx_with_merged_header_auto_recovers(self, tmp_path):
        """XLSX com título mergeado na linha 1 deve auto-detectar header_row=2
        e retornar schema correto (não 1 coluna VARCHAR única)."""
        xlsx = tmp_path / "dados.xlsx"
        _build_patho_xlsx(xlsx)
        result = asyncio.run(analyze_tabular(xlsx.read_bytes(), "dados.xlsx"))

        assert result["sheet_count"] == 2
        assert result["primary_sheet"] in ("TB_ANALISE_CREDITO", "TB_PREDICOES")
        # Top-level reflete aba primária, score deve ser alto após auto-detect
        assert result["score"] >= 0.7, f"score baixo: {result}"
        assert result["tabular_ready"] is True

        # Cada aba deve ter sido recuperada (não 1 coluna patológica)
        for s in result["sheets"]:
            assert s["columns"] >= 3, f"aba {s['name']} ficou com {s['columns']} cols"
            assert s["header_row"] == 2, f"esperava header_row=2 em {s['name']}"
            assert s["header_row_auto_detected"] is True
            assert s["tabular_ready"] is True
            # PK auto-detectada
            assert s["suggested_pk"] == "cd_cliente"

    def test_xlsx_lists_all_sheets(self, tmp_path):
        """sheets[] deve listar TODAS as abas, na ordem do arquivo."""
        xlsx = tmp_path / "dados.xlsx"
        _build_patho_xlsx(xlsx)
        result = asyncio.run(analyze_tabular(xlsx.read_bytes(), "dados.xlsx"))
        names = [s["name"] for s in result["sheets"]]
        assert names == ["TB_ANALISE_CREDITO", "TB_PREDICOES"]

    def test_csv_keeps_legacy_format(self):
        """CSV mantém top-level com schema (backward compat) + sheets=[1 entry]."""
        result = asyncio.run(analyze_tabular(CSV_OK, "test.csv"))
        # sheet_count=1 (CSV é 1 aba virtual)
        assert result["sheet_count"] == 1
        # primary_sheet é None pra CSV (não tem nome de aba)
        assert result.get("primary_sheet") is None
        # Top-level continua disponível
        assert result["rows"] == 4
        assert result["columns"] == 4


class TestPatologicHeuristic:
    """Casos onde a heurística DEVE penalizar (sem usar XLSX real)."""

    def test_title_like_column_name_penalized(self):
        """Coluna nomeada como UPPER_SNAKE_CASE longo deve disparar penalidade
        mesmo sem ser caso patológico extremo."""
        from app.evidence.tabular import _looks_like_table_title
        assert _looks_like_table_title("TB_ANALISE_CREDITO") is True
        assert _looks_like_table_title("FATO_VENDAS_MENSAL") is True
        assert _looks_like_table_title("DIM_CLIENTE") is True
        assert _looks_like_table_title("nome_coluna_normal") is False
        assert _looks_like_table_title("ID") is False  # curto, sem _


# ═══════════════════════════════════════════════════════════════
# 4. promote_to_table — DuckDB real + Postgres mockado
# ═══════════════════════════════════════════════════════════════


@pytest.fixture
def isolated_storage(monkeypatch, tmp_path):
    """Isola _TABULAR_ROOT em tmp_path + mocka repos Postgres + queries.

    Retorna dict com handles para inspecionar/manipular o estado.
    """
    monkeypatch.setattr(tabular_service, "_TABULAR_ROOT", tmp_path / "tabular")

    # KS mock
    ks_store = {"ks-1": {
        "id": "ks-1",
        "name": "Vendas",
        "confidentiality_label": "internal",
        "authorized": 1,
    }}

    async def fake_ks_find(ks_id):
        return ks_store.get(ks_id)

    monkeypatch.setattr(tabular_service.knowledge_repo, "find_by_id", fake_ks_find)

    # data_tables_repo mock
    table_store: dict[str, dict] = {}

    async def fake_table_create(data):
        table_store[data["id"]] = dict(data)
        return data

    monkeypatch.setattr(tabular_service.data_tables_repo, "create", fake_table_create)

    # data_table_query_logs_repo mock (best-effort, audit não derruba)
    audit_log: list[dict] = []

    async def fake_audit_create(data):
        audit_log.append(dict(data))
        return data

    monkeypatch.setattr(tabular_service.data_table_query_logs_repo, "create", fake_audit_create)

    # find_by_id_with_ks: monta enriquecido a partir de table_store + ks_store
    async def fake_find_by_id(table_id):
        t = table_store.get(table_id)
        if not t:
            return None
        ks = ks_store.get(t["knowledge_source_id"], {})
        out = dict(t)
        out["ks_confidentiality_label"] = ks.get("confidentiality_label")
        out["ks_authorized"] = ks.get("authorized")
        out["ks_name"] = ks.get("name")
        return out

    monkeypatch.setattr(tabular_service, "find_by_id_with_ks", fake_find_by_id)

    # next_version_for_slug: começa em 1, incrementa por slug
    async def fake_next_version(ks_id, slug):
        existing = [t for t in table_store.values()
                    if t["knowledge_source_id"] == ks_id
                    and t.get("urn", "").endswith(f":{slug}:" + t.get("version", ""))]
        # Conta tabelas com mesmo slug
        same_slug = [t for t in table_store.values()
                     if t["knowledge_source_id"] == ks_id
                     and f":{slug}:" in t.get("urn", "")]
        return len(same_slug) + 1

    monkeypatch.setattr(tabular_service, "next_version_for_slug", fake_next_version)

    return {
        "ks_store": ks_store,
        "table_store": table_store,
        "audit_log": audit_log,
        "tmp_root": tmp_path / "tabular",
    }


class TestPromoteToTable:
    def test_promote_csv_creates_duckdb_file(self, isolated_storage):
        result = asyncio.run(promote_to_table(
            ks_id="ks-1", data=CSV_OK, filename="vendas.csv", name="Vendas Q4",
        ))
        assert result["status"] == "ready"
        assert result["row_count"] == 4
        assert result["column_count"] == 4
        # Arquivo .duckdb foi criado
        path = Path(result["duckdb_path"])
        assert path.exists()
        assert path.stat().st_size > 0

    def test_promote_unknown_ks_raises_404(self, isolated_storage):
        with pytest.raises(TabularError) as exc:
            asyncio.run(promote_to_table(
                ks_id="nope", data=CSV_OK, filename="x.csv",
            ))
        assert exc.value.status_code == 404

    def test_promote_versions_increment_on_same_slug(self, isolated_storage):
        r1 = asyncio.run(promote_to_table(
            ks_id="ks-1", data=CSV_OK, filename="vendas.csv", name="Vendas Q4",
        ))
        r2 = asyncio.run(promote_to_table(
            ks_id="ks-1", data=CSV_OK, filename="vendas.csv", name="Vendas Q4",
        ))
        assert r1["urn"] != r2["urn"]
        assert ":1" in r1["urn"]
        assert ":2" in r2["urn"]

    def test_promote_xlsx_specific_sheet_creates_one_table(self, isolated_storage, tmp_path):
        """XLSX com 2 abas + sheet_name='TB_PREDICOES' → cria 1 tabela só
        daquela aba, com display name diferenciado."""
        xlsx = tmp_path / "dados.xlsx"
        _build_patho_xlsx(xlsx)
        result = asyncio.run(promote_to_table(
            ks_id="ks-1", data=xlsx.read_bytes(), filename="dados.xlsx",
            sheet_name="TB_PREDICOES",
        ))
        assert result["status"] == "ready"
        # Display name contém o sheet_name pra diferenciar
        assert "TB_PREDICOES" in result["name"]
        # URN incorpora o sheet_name (slug com __ separador)
        assert "predicoes" in result["urn"].lower()
        # Schema correto (3 colunas — cd_cliente, score, decisao)
        assert result["column_count"] == 3
        assert result["row_count"] == 10

    def test_promote_xlsx_unknown_sheet_raises_400(self, isolated_storage, tmp_path):
        xlsx = tmp_path / "dados.xlsx"
        _build_patho_xlsx(xlsx)
        with pytest.raises(TabularError) as exc:
            asyncio.run(promote_to_table(
                ks_id="ks-1", data=xlsx.read_bytes(), filename="dados.xlsx",
                sheet_name="ABA_INEXISTENTE",
            ))
        assert exc.value.status_code == 400
        assert "ABA_INEXISTENTE" in str(exc.value)

    def test_promote_xlsx_default_sheet_uses_primary(self, isolated_storage, tmp_path):
        """Sem sheet_name, usa a aba primária da análise (top-level)."""
        xlsx = tmp_path / "dados.xlsx"
        _build_patho_xlsx(xlsx)
        result = asyncio.run(promote_to_table(
            ks_id="ks-1", data=xlsx.read_bytes(), filename="dados.xlsx",
        ))
        assert result["status"] == "ready"
        # 4 ou 3 cols (depende de qual aba teve maior score — TB_ANALISE tem 4)
        assert result["column_count"] in (3, 4)


# ═══════════════════════════════════════════════════════════════
# 5. execute_query — DuckDB real, bind vars, read-only
# ═══════════════════════════════════════════════════════════════


class TestExecuteQuery:
    @pytest.fixture
    def ready_table(self, isolated_storage):
        """Promove CSV_OK e retorna metadata da tabela criada."""
        return asyncio.run(promote_to_table(
            ks_id="ks-1", data=CSV_OK, filename="vendas.csv", name="Vendas",
        ))

    def test_select_all_returns_all_rows(self, ready_table):
        r = asyncio.run(execute_query(table_id=ready_table["id"]))
        assert r["row_count"] == 4
        assert "id" in r["columns"]

    def test_select_specific_columns(self, ready_table):
        r = asyncio.run(execute_query(
            table_id=ready_table["id"],
            select=["nome", "idade"],
        ))
        assert r["columns"] == ["nome", "idade"]
        assert all(set(row.keys()) == {"nome", "idade"} for row in r["rows"])

    def test_where_eq_filter(self, ready_table):
        r = asyncio.run(execute_query(
            table_id=ready_table["id"],
            filters=[{"col": "nome", "op": "=", "value": "Alice"}],
        ))
        assert r["row_count"] == 1
        assert r["rows"][0]["nome"] == "Alice"

    def test_where_between_filter(self, ready_table):
        r = asyncio.run(execute_query(
            table_id=ready_table["id"],
            filters=[{"col": "idade", "op": "BETWEEN", "value": [25, 35]}],
        ))
        # Alice 30, Bob 25, Dan 33 — Carol 42 out
        assert r["row_count"] == 3

    def test_where_in_filter(self, ready_table):
        r = asyncio.run(execute_query(
            table_id=ready_table["id"],
            filters=[{"col": "nome", "op": "IN", "value": ["Alice", "Carol"]}],
        ))
        assert r["row_count"] == 2

    def test_where_is_null_filter(self, isolated_storage):
        promoted = asyncio.run(promote_to_table(
            ks_id="ks-1", data=CSV_WITH_NULLS, filename="nulls.csv", name="N",
        ))
        r = asyncio.run(execute_query(
            table_id=promoted["id"],
            filters=[{"col": "email", "op": "IS NULL"}],
        ))
        # Alice e Bob têm email vazio → DuckDB read_csv_auto trata "" como NULL
        assert r["row_count"] == 2

    def test_template_input_resolves(self, ready_table):
        r = asyncio.run(execute_query(
            table_id=ready_table["id"],
            inputs={"target_name": "Bob"},
            filters=[{"col": "nome", "op": "=", "value": "{{ inputs.target_name }}"}],
        ))
        assert r["row_count"] == 1
        assert r["rows"][0]["nome"] == "Bob"

    def test_if_present_skips_when_input_missing(self, ready_table):
        r = asyncio.run(execute_query(
            table_id=ready_table["id"],
            inputs={},  # sem target_name
            filters=[{
                "col": "nome", "op": "=", "value": "{{ inputs.target_name }}",
                "if_present": "target_name",
            }],
        ))
        # Filtro pulado → todas as linhas
        assert r["row_count"] == 4

    def test_order_by_and_limit(self, ready_table):
        r = asyncio.run(execute_query(
            table_id=ready_table["id"],
            order_by=["idade DESC"],
            limit=2,
        ))
        assert r["row_count"] == 2
        # Carol(42) e Dan(33) primeiro
        assert r["rows"][0]["nome"] == "Carol"

    def test_unknown_column_in_select_raises_400(self, ready_table):
        with pytest.raises(TabularError) as exc:
            asyncio.run(execute_query(
                table_id=ready_table["id"],
                select=["coluna_inexistente"],
            ))
        assert exc.value.status_code == 400

    def test_unknown_operator_raises_400(self, ready_table):
        with pytest.raises(TabularError) as exc:
            asyncio.run(execute_query(
                table_id=ready_table["id"],
                filters=[{"col": "nome", "op": "DROP TABLE", "value": "x"}],
            ))
        assert exc.value.status_code == 400

    def test_limit_above_cap_rejected(self, ready_table):
        with pytest.raises(TabularError) as exc:
            asyncio.run(execute_query(
                table_id=ready_table["id"],
                limit=MAX_ROWS_RETURNED + 1,
            ))
        assert exc.value.status_code == 400

    def test_unknown_table_raises_404(self, isolated_storage):
        with pytest.raises(TabularError) as exc:
            asyncio.run(execute_query(table_id="nope-uuid"))
        assert exc.value.status_code == 404

    def test_read_only_enforcement_blocks_writes(self, ready_table):
        """DuckDB read_only=True deve barrar INSERT/UPDATE/DELETE/DROP.

        Testamos diretamente abrindo a conexão como o engine faz:
        a tentativa de modificar deve estourar exception.
        """
        path = ready_table["duckdb_path"]
        con = duckdb.connect(path, read_only=True)
        try:
            with pytest.raises(Exception):
                con.execute("INSERT INTO data VALUES (99, 'x', 99, 0)")
            with pytest.raises(Exception):
                con.execute("DROP TABLE data")
        finally:
            con.close()


# ═══════════════════════════════════════════════════════════════
# 6. Routes — mini FastAPI + mocks
# ═══════════════════════════════════════════════════════════════


def make_app(user: dict) -> FastAPI:
    app = FastAPI()
    app.include_router(data_tables_router)
    app.dependency_overrides[require_user] = lambda: user
    return app


@pytest.fixture
def root_user():
    return {"id": "u-root", "role": "root", "domains": "[]"}


@pytest.fixture
def common_user():
    return {"id": "u-comum", "role": "comum", "domains": "[]"}


class TestRoutes:
    def test_get_unknown_table_returns_404(self, monkeypatch, root_user):
        async def fake_find(_id):
            return None

        monkeypatch.setattr("app.routes.data_tables.find_by_id_with_ks", fake_find)
        client = TestClient(make_app(root_user))
        r = client.get("/api/v1/data-tables/nope")
        assert r.status_code == 404

    def test_get_table_blocked_by_visibility_returns_403(self, monkeypatch, common_user):
        async def fake_find(_id):
            return {
                "id": _id,
                "ks_authorized": 1,
                "ks_confidentiality_label": "restricted",
            }

        monkeypatch.setattr("app.routes.data_tables.find_by_id_with_ks", fake_find)
        client = TestClient(make_app(common_user))
        r = client.get("/api/v1/data-tables/t-1")
        assert r.status_code == 403

    def test_get_table_root_bypasses_visibility(self, monkeypatch, root_user):
        async def fake_find(_id):
            return {
                "id": _id,
                "ks_authorized": 0,
                "ks_confidentiality_label": "confidential",
                "name": "Sigiloso",
            }

        monkeypatch.setattr("app.routes.data_tables.find_by_id_with_ks", fake_find)
        client = TestClient(make_app(root_user))
        r = client.get("/api/v1/data-tables/t-1")
        assert r.status_code == 200
        assert r.json()["name"] == "Sigiloso"

    def test_list_data_tables_returns_envelope(self, monkeypatch, root_user):
        async def fake_list(user, ks_id=None):
            return [{"id": "t1", "name": "T1"}, {"id": "t2", "name": "T2"}]

        monkeypatch.setattr("app.routes.data_tables.list_for_user", fake_list)
        client = TestClient(make_app(root_user))
        r = client.get("/api/v1/data-tables")
        assert r.status_code == 200
        body = r.json()
        assert body["total"] == 2
        assert len(body["data_tables"]) == 2

    def test_query_endpoint_404_when_table_missing(self, monkeypatch, root_user):
        async def fake_find(_id):
            return None

        monkeypatch.setattr("app.routes.data_tables.find_by_id_with_ks", fake_find)
        client = TestClient(make_app(root_user))
        r = client.post("/api/v1/data-tables/nope/query", json={})
        assert r.status_code == 404

    def test_query_endpoint_403_when_visibility_blocks(self, monkeypatch, common_user):
        async def fake_find(_id):
            return {
                "id": _id,
                "ks_authorized": 1,
                "ks_confidentiality_label": "confidential",
            }

        monkeypatch.setattr("app.routes.data_tables.find_by_id_with_ks", fake_find)
        client = TestClient(make_app(common_user))
        r = client.post("/api/v1/data-tables/t-1/query", json={})
        assert r.status_code == 403
